#-------------------------------------------------------------------------------
# Name:        refpExcel
# Purpose:	   Auxiliary functions to import/export from/to MS Excel files
#
# Author:      roberto ferrer
#
# Created:     31/05/2012
# Copyright:   (c) roberto ferrer 2012
# Licence:     GNU General Public License
#-------------------------------------------------------------------------------
#!/usr/bin/env python

def main():
    pass

if __name__ == '__main__':
    main()


import sys
import numpy as np
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter



def openExcel (fileDir,sheetName,numRange):
    #reading excel files (.XLSM) using openpyxl
    fi = fileDir
    sh = sheetName
    nr = numRange

    print 'Loading Excel file...'
    wb = load_workbook(fi)
    ws = wb.get_sheet_by_name(sh)
    #get numbers
    matm = ws.range(nr)
    matmlist = []
    row = range(len(matm)) #number of 'rows' imported to matm tuple
    column = range(len(matm[0]))  #number of 'columns' imported to matm tuple

    print 'Fixing formats for',sh,'...'
    for i in row:
        #process numbers
        nvalues = [matm[i][j].value for j in column]  #traverse columns while keeping row fixed
        nvalues = [0 if x==None else x for x in nvalues] #convert none type to zeros (happens when Excel cell is empty).
        matmlist.append(nvalues)
    #convert matrices to array
    matarr = np.array(matmlist)

    print sh,'loaded.'
    print 'Matrix has shape',matarr.shape,'\n'

    return matarr


def sendExcel(fileDir,matrixList,sheetNameList):
    #check that input lists have same length (one sheet name for each matrix)
    if len(matrixList) != len(sheetNameList):
        print 'List of variables being exported and of sheet names must be equal.'
        return

    #Set up excel workbook
    wb = Workbook()
    ew = ExcelWriter(workbook = wb)
    dest_filename = fileDir
    #Create sheet and export data for every matrix being exported
    for matrix,sheetN in zip(matrixList,sheetNameList):
        ws = wb.create_sheet()
        ws.title = sheetN
        #For a matrix
        if len(matrix.shape) == 2:
            #Dump matrix in the Excel sheet
            (r,c) = matrix.shape
            for element in xrange(c):
                col = get_column_letter(element+2) #2, so it will start printing on B
                for row in xrange(r):
                    ws.cell('%s%s'%(col, row+2)).value = matrix[row][element]
        #For a vector
        elif len(matrix.shape) == 1:
            r = matrix.shape
            r = r[0]
            col = get_column_letter(2) #2, so it will start printing on B
            for row in xrange(r):
                ws.cell('%s%s'%(col, row+2)).value = matrix[row]
        else:
            print 'Only exports matrices and vectors!'
            print 'Weird element found. \n'
            return

##    #Create another sheet in same workbook and put Pi
##    ws = wb.create_sheet()
##    ws.title = 'Pi'
##    ws.cell('F5').value = 3.14

    ew.save(dest_filename)
    print 'Matrix has been exported to Excel.'


def sendExcel2(fileDir, iterableObject, sheetName):
    ''' Export objects in an iterable (e.g. List) to a column in Excel.
    Everything is exported as text.
    For example, a list of tuples can be exported to a column in Excel.
    ToDo: Probably better to expand the original sendExcel function so it can
    handle various types of data (e.g. numpy matrices, vectors, iterables, etc.)
    '''
    #Set up excel workbook
    wb = Workbook()
    ew = ExcelWriter(workbook = wb)
    dest_filename = fileDir

    ws = wb.create_sheet()
    ws.title = sheetName

    r = len(iterableObject)
    col = get_column_letter(2) #2, so it will start printing on B
    for row in xrange(r):
        s = str(iterableObject[row])
        ws.cell('%s%s'%(col, row+2)).value = s
    wb.save(filename = dest_filename)
    print 'Objects have been exported to Excel.'